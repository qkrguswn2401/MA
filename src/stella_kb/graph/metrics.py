"""Cell -> Metric lift: name the raw cells.

`extract.py`/`graph.py` give a cell DAG and Section/Sheet/Fund/Entity nodes, but the
financial line items (AUM, fees, EV, WACC, FCFF) still live only as raw ``Sheet!Ref``
cells. This module attaches **named `Metric` nodes** to the specific cells that hold them,
turning "DCF!K59 depends on …" into "Equity Value <-driven by- AUM".

It is a **curated anchor table** (`METRICS`), not an LLM guess: every metric points at an
exact, verified cell/row. This is the OpenKB whitelist pattern — `METRIC_IDS` is the closed
vocabulary, and cross-metric edges (`DRIVES`/`ASSUMPTION_OF`) may only connect ids that
exist in it, so no edge can reference a metric that was never defined.

Anchor kinds:
- ``scalar``  one cell -> one Metric (value on the node); optional ``period`` for dated items.
- ``series``  a row on a sheet with a "Fiscal Year" header -> one ``HAS_VALUE`` edge per year.
- ``vseries`` years run down a column (AUM, headcount) -> ``HAS_VALUE`` per row.

Edges emitted: ``DEFINED_IN`` (Metric -> Cell and Metric -> Sheet), ``HAS_VALUE``
(Metric -> Period, value on the edge), ``DRIVES`` and ``ASSUMPTION_OF`` (Metric -> Metric).
"""

from __future__ import annotations

from dataclasses import dataclass, field

import networkx as nx
import openpyxl
from openpyxl.utils import get_column_letter


# --- metric registry (the closed vocabulary) ------------------------------------------

@dataclass
class Metric:
    id: str
    label_en: str
    sheet: str
    category: str                 # valuation | assumption | revenue | expense | cashflow | driver | bridge
    kind: str                     # scalar | series | vseries
    cell: str | None = None       # scalar: "K59"
    row: int | None = None        # series: row number; horizontal years from the Fiscal Year header
    year_col: str | None = None   # vseries: column holding the year (e.g. "C")
    val_col: str | None = None    # vseries: column holding the value (e.g. "F")
    row_range: tuple[int, int] | None = None  # vseries: inclusive row span
    period: int | None = None     # scalar: the fiscal year this value is dated to, if any
    case: str | None = None       # "DTT" | "MGT" — which valuation case the cell reflects
    label_ko: str | None = None
    aliases: list[str] = field(default_factory=list)


# Active case in the file is DTT (DCF!J6 == 2); the DCF summary cells reflect it.
METRICS: list[Metric] = [
    # -- valuation outputs (DCF summary bridge) --
    Metric("equity_value", "Equity Value", "DCF", "valuation", "scalar", cell="K59", case="DTT",
           label_ko="지분가치", aliases=["equity value", "주주가치"]),
    Metric("enterprise_value", "Enterprise Value", "DCF", "valuation", "scalar", cell="K57", case="DTT",
           label_ko="기업가치", aliases=["EV", "enterprise value"]),
    Metric("operating_value", "Operating Value", "DCF", "valuation", "scalar", cell="K55", case="DTT",
           label_ko="영업가치"),
    Metric("pv_projection_fcf", "PV of Projection-Period FCF", "DCF", "valuation", "scalar", cell="K53", case="DTT"),
    Metric("pv_terminal_value", "PV of Terminal Value", "DCF", "valuation", "scalar", cell="K54", case="DTT"),
    Metric("noa", "Non-Operating Assets", "DCF", "valuation", "scalar", cell="K56", case="DTT",
           label_ko="비영업자산"),
    Metric("net_cash_debt", "Net Cash/(Debt)", "DCF", "valuation", "scalar", cell="K58", case="DTT",
           label_ko="순현금/(순차입금)"),

    # -- assumptions --
    Metric("wacc", "WACC (discount rate)", "DCF", "assumption", "scalar", cell="K45",
           label_ko="할인율", aliases=["discount rate", "할인율"]),
    Metric("perpetual_growth_rate", "Perpetual Growth Rate", "DCF", "assumption", "scalar", cell="K46",
           label_ko="영구성장률", aliases=["PGR", "terminal growth"]),
    Metric("valuation_date", "Valuation Date", "DCF", "assumption", "scalar", cell="K44",
           label_ko="평가기준일"),
    Metric("hurdle_rate", "Hurdle Rate", "IRR", "assumption", "scalar", cell="W13",
           label_ko="기준수익률", aliases=["preferred return", "기준수익률"]),
    Metric("carry_rate", "Carried-Interest Rate", "IRR", "assumption", "scalar", cell="W14",
           label_ko="초과수익률", aliases=["carry", "performance fee rate", "초과수익률"]),

    # -- operating revenue lines (Operating Revenue, horizontal FY series) --
    Metric("operating_revenue_total", "Total Operating Revenue", "Operating Revenue", "revenue", "series", row=43,
           label_ko="영업수익"),
    Metric("management_fee", "Management Fee", "Operating Revenue", "revenue", "series", row=38,
           label_ko="관리보수", aliases=["관리수수료", "mgmt fee", "mgmt revenue"]),
    Metric("performance_fee", "Performance Fee (carry)", "Operating Revenue", "revenue", "series", row=39,
           label_ko="성과보수", aliases=["carried interest", "carry"]),
    Metric("dividend_income", "Dividend Income", "Operating Revenue", "revenue", "series", row=40,
           label_ko="배당금", aliases=["배당금수익", "distributions"]),
    Metric("advisory_fee", "Advisory Service Fee", "Operating Revenue", "revenue", "series", row=41,
           label_ko="자문용역수수료"),

    # -- operating expense lines (Operating Expense, horizontal FY series) --
    Metric("operating_expense_total", "Total Operating Expense", "Operating Expense", "expense", "series", row=42,
           label_ko="영업비용"),
    Metric("personnel_cost", "Personnel Cost", "Operating Expense", "expense", "series", row=38,
           label_ko="인건비", aliases=["labor cost", "comp"]),
    Metric("other_expense", "Other Expense", "Operating Expense", "expense", "series", row=39,
           label_ko="기타비용"),
    Metric("depreciation_amortization", "Depreciation & Amortization", "Operating Expense", "expense", "series", row=40,
           label_ko="D&A"),

    # -- DCF cash-flow build (DCF sheet, horizontal FY series) --
    Metric("ebit", "EBIT", "DCF", "cashflow", "series", row=25, label_ko="영업이익"),
    Metric("ebitda", "EBITDA", "DCF", "cashflow", "series", row=28),
    Metric("change_nwc", "Change in NWC", "DCF", "cashflow", "series", row=30, label_ko="순운전자본증감"),
    Metric("capex", "CapEx", "DCF", "cashflow", "series", row=31, label_ko="자본적지출"),
    Metric("gp_commitment_cf", "GP Commitment (cash)", "DCF", "cashflow", "series", row=32, label_ko="GP출자"),
    Metric("tax_cf", "Cash Tax", "DCF", "cashflow", "series", row=33, label_ko="법인세"),
    Metric("fcff", "Free Cash Flow to Firm", "DCF", "cashflow", "series", row=34),

    # -- EV -> equity bridge components (Net debt, NOA — dated to FY24) --
    Metric("net_debt_total", "Net Debt (total)", "Net debt, NOA", "bridge", "scalar", cell="R36", period=2024,
           label_ko="순차입금"),
    Metric("cash", "Cash & Equivalents", "Net debt, NOA", "bridge", "scalar", cell="R37", period=2024,
           label_ko="현금및현금성자산"),
    Metric("short_term_debt", "Short-Term Borrowings", "Net debt, NOA", "bridge", "scalar", cell="R38", period=2024,
           label_ko="단기차입금"),
    Metric("current_lt_debt", "Current Portion of LT Debt", "Net debt, NOA", "bridge", "scalar", cell="R39", period=2024,
           label_ko="유동성장기부채"),
    Metric("severance_provision", "Severance Provision", "Net debt, NOA", "bridge", "scalar", cell="R40", period=2024,
           label_ko="퇴직급여충당부채"),
    Metric("noa_total", "Non-Operating Assets (total)", "Net debt, NOA", "bridge", "scalar", cell="R42", period=2024,
           label_ko="비영업자산"),

    # -- drivers (vertical series: years down a column) --
    Metric("aum_cumulative", "Cumulative AUM", "AUM Projection", "driver", "vseries",
           year_col="C", val_col="F", row_range=(5, 16),
           label_ko="누적 AUM", aliases=["AUM", "assets under management"]),
    Metric("headcount", "Headcount", "인력", "driver", "vseries",
           year_col=None, val_col=None, row_range=(38, 38),  # special: total row, years across D..I
           label_ko="임직원 수", aliases=["FTE", "employees"]),
]

METRIC_IDS = {m.id for m in METRICS}


# Dual-case: the live `DCF` sheet (and thus each metric's primary cell) holds whichever case
# is active — currently **DTT**. The **MGT** counterpart is frozen in the MGT exhibit, which
# shares the DTT exhibit's exact layout, so each DCF-summary metric's MGT value sits at a known
# cell there. We keep the metric's DEFINED_IN anchor on the live DCF cell (it wires into the
# formula DAG; the exhibit is a downstream PPT view, never the source of truth) and attach the
# MGT figure as `value_mgt` + `cell_mgt`. So `value`=DTT (active), `value_mgt`=MGT.
MGT_EXHIBIT = "DCF 장표 #1_MGT"
DUAL_CASE_MGT: dict[str, str] = {
    "equity_value": "E12", "enterprise_value": "E10", "operating_value": "E8",
    "pv_projection_fcf": "E6", "pv_terminal_value": "E7", "noa": "E9",
    "net_cash_debt": "E11", "wacc": "J5", "perpetual_growth_rate": "J6",
    "valuation_date": "J7",
}


# Per-fund fee anchors live in `관리수수료` rows 8-19 (name=C, committed=O, rate=P, annual=Q).
# Map each fee-sheet fund label to the Biz Plan `Fund:` node id, where one exists (the fee
# sheet uses "제1호 차이나" where Biz Plan uses "차이나1호"; some funds have no Biz Plan sheet).
FUND_FEE_ROWS = (8, 19)
FUND_NODE_MAP: dict[str, str] = {
    "제1호 차이나": "차이나1호",
    "제2호 바이아웃": "제2호",
    "제3호 그로쓰": "제3호",
    "옐로씨 제1호": "옐로씨",
    "제5호 바이아웃": "제5호",
    "제7호 바이아웃": "7호&7-1호",
    "제7-1호 바이아웃": "7호&7-1호",
    "제8호 코인베스트": "제8호",
    # PCC신기술 / 웰릭스신기술 / 제4호 바이아웃 / 기타 — no Biz Plan fund node
}


# Per-fund GP **carry** anchors. The `성과보수, 배당금` sheet lays six fund blocks side by
# side, each with a *different* value-column offset, so it defeats the generic series reader —
# hence a curated cell table (the same anchor philosophy as METRICS / FUND_FEE_ROWS). In every
# block the headline rows are 성과보수 (carry) MGT=row4 / DTT=row6 and 재산분배액
# (distribution) MGT=row7 / DTT=row9; the Exit assumptions sit on row 5. The active file case
# is DTT, so `value` is the DTT figure and `value_mgt` carries the management-case figure.
# `node` is the Biz Plan `Fund:` node id (제7호/제7-1호 share the combined `7호&7-1호` fund).
CARRY_FUNDS: list[dict] = [
    {"alias": "제2호", "node": "제2호", "val": "E", "ebitda": "J5", "mult": "L5", "hurdle": "N5"},
    {"alias": "옐로씨", "node": "옐로씨", "val": "S", "ebitda": "V5", "mult": "W5", "hurdle": "Y5"},
    {"alias": "제5호", "node": "제5호", "val": "AF", "ebitda": "AJ5", "mult": "AK5", "hurdle": "AM5"},
    {"alias": "제7호", "node": "7호&7-1호", "val": "AU", "ebitda": "AY5", "mult": "AZ5", "hurdle": "BB5"},
    {"alias": "제7-1호", "node": "7호&7-1호", "val": "BI", "ebitda": "BM5", "mult": "BN5", "hurdle": "BP5"},
    {"alias": "제8호", "node": "제8호", "val": "BW", "ebitda": "CB5", "mult": "CC5", "hurdle": "CE5"},
]
CARRY_SHEET = "성과보수, 배당금"


def _slug(name: str) -> str:
    return name.strip().replace(" ", "").replace("-", "")


# cross-metric edges — endpoints MUST be in METRIC_IDS (whitelist guard applied below)
DRIVES: list[tuple[str, str]] = [
    ("aum_cumulative", "management_fee"),
    ("aum_cumulative", "performance_fee"),
    ("management_fee", "operating_revenue_total"),
    ("performance_fee", "operating_revenue_total"),
    ("dividend_income", "operating_revenue_total"),
    ("advisory_fee", "operating_revenue_total"),
    ("operating_revenue_total", "ebit"),
    ("operating_expense_total", "ebit"),
    ("personnel_cost", "operating_expense_total"),
    ("other_expense", "operating_expense_total"),
    ("depreciation_amortization", "operating_expense_total"),
    ("ebit", "ebitda"),
    ("ebitda", "fcff"),
    ("change_nwc", "fcff"),
    ("capex", "fcff"),
    ("gp_commitment_cf", "fcff"),
    ("tax_cf", "fcff"),
    ("fcff", "pv_projection_fcf"),
    ("pv_projection_fcf", "operating_value"),
    ("pv_terminal_value", "operating_value"),
    ("operating_value", "enterprise_value"),
    ("noa", "enterprise_value"),
    ("noa_total", "noa"),
    ("enterprise_value", "equity_value"),
    ("net_cash_debt", "equity_value"),
    ("net_debt_total", "net_cash_debt"),
    ("cash", "net_debt_total"),
    ("short_term_debt", "net_debt_total"),
    ("current_lt_debt", "net_debt_total"),
    ("severance_provision", "net_debt_total"),
    ("headcount", "personnel_cost"),
]
ASSUMPTION_OF: list[tuple[str, str]] = [
    ("wacc", "pv_projection_fcf"),
    ("wacc", "equity_value"),
    ("perpetual_growth_rate", "pv_terminal_value"),
    ("valuation_date", "equity_value"),
    ("hurdle_rate", "performance_fee"),
    ("carry_rate", "performance_fee"),
]


# --- period resolution ----------------------------------------------------------------

def fiscal_year_axis(ws) -> dict[str, object]:
    """Find the ``Fiscal Year`` header row and map each column letter -> year (or 'TV').

    Different sheets offset the year columns (DCF starts a column earlier than the engine
    sheets), so the axis is read per sheet rather than hardcoded.
    """
    for r in range(1, 21):
        for c in range(1, 41):
            if ws.cell(row=r, column=c).value == "Fiscal Year":
                axis: dict[str, object] = {}
                for c2 in range(c + 1, c + 16):
                    v = ws.cell(row=r, column=c2).value
                    if v is None:
                        continue
                    axis[get_column_letter(c2)] = "TV" if isinstance(v, str) else int(v)
                return axis
    return {}


def _period_id(year: object) -> str:
    return f"Period:{year}"


# --- graph construction ---------------------------------------------------------------

def attach_metrics(g: nx.DiGraph, path: str) -> nx.DiGraph:
    """Add Metric/Period nodes and their edges onto an existing semantic graph ``g``."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    axis_cache: dict[str, dict[str, object]] = {}

    def add_period(year: object) -> str:
        pid = _period_id(year)
        if pid not in g:
            g.add_node(pid, type="Period", year=year)
        return pid

    for m in METRICS:
        mid = f"Metric:{m.id}"
        g.add_node(mid, type="Metric", label=m.label_en, label_ko=m.label_ko,
                   category=m.category, sheet=m.sheet, case=m.case, aliases=m.aliases)
        # tie the metric to its sheet node if the semantic graph has one
        sheet_node = f"Sheet:{m.sheet}"
        if sheet_node in g:
            g.add_edge(mid, sheet_node, type="DEFINED_IN")

        ws = wb[m.sheet]
        if m.kind == "scalar":
            cid = f"{m.sheet}!{m.cell}"
            val = ws[m.cell].value
            g.nodes[mid]["value"] = val
            g.add_node(cid, type="Cell", sheet=m.sheet)
            g.add_edge(mid, cid, type="DEFINED_IN")
            if m.id in DUAL_CASE_MGT:                  # add the frozen MGT-case counterpart
                ec = DUAL_CASE_MGT[m.id]
                g.nodes[mid]["value_mgt"] = wb[MGT_EXHIBIT][ec].value
                g.nodes[mid]["cell_mgt"] = f"{MGT_EXHIBIT}!{ec}"
            if m.period is not None:
                pid = add_period(m.period)
                g.add_edge(mid, pid, type="HAS_VALUE", value=val, cell=cid)

        elif m.kind == "series":
            axis = axis_cache.setdefault(m.sheet, fiscal_year_axis(ws))
            for col, year in axis.items():
                cell = f"{col}{m.row}"
                val = ws[cell].value
                if val is None:
                    continue
                pid = add_period(year)
                cid = f"{m.sheet}!{cell}"
                g.add_node(cid, type="Cell", sheet=m.sheet)
                g.add_edge(mid, pid, type="HAS_VALUE", value=val, cell=cid)
                g.add_edge(mid, cid, type="DEFINED_IN")

        elif m.kind == "vseries":
            r0, r1 = m.row_range
            if m.year_col and m.val_col:  # years down a column (AUM)
                for r in range(r0, r1 + 1):
                    yraw = ws[f"{m.year_col}{r}"].value
                    val = ws[f"{m.val_col}{r}"].value
                    if yraw is None or val is None:
                        continue
                    year = _year_of(yraw)
                    pid = add_period(year)
                    cell = f"{m.val_col}{r}"
                    cid = f"{m.sheet}!{cell}"
                    g.add_node(cid, type="Cell", sheet=m.sheet)
                    g.add_edge(mid, pid, type="HAS_VALUE", value=val, cell=cid)
                    g.add_edge(mid, cid, type="DEFINED_IN")
            else:  # headcount: a total row with years across the header above it
                hdr = {get_column_letter(c): ws.cell(row=r0 - 32, column=c).value
                       for c in range(4, 10)}  # row 6 holds '2019년'… for the row-38 total
                for c in range(4, 10):
                    val = ws.cell(row=r0, column=c).value
                    yraw = hdr.get(get_column_letter(c))
                    if val is None or yraw is None:
                        continue
                    year = _year_of(yraw)
                    cell = f"{get_column_letter(c)}{r0}"
                    cid = f"{m.sheet}!{cell}"
                    pid = add_period(year)
                    g.add_node(cid, type="Cell", sheet=m.sheet)
                    g.add_edge(mid, pid, type="HAS_VALUE", value=val, cell=cid)
                    g.add_edge(mid, cid, type="DEFINED_IN")

    # -- per-fund management-fee anchors (관리수수료 rows 8-19) --
    fee_ws = wb["관리수수료"]
    r0, r1 = FUND_FEE_ROWS
    for r in range(r0, r1 + 1):
        name = fee_ws.cell(row=r, column=3).value          # C: fund name
        if not name:
            continue
        committed = fee_ws[f"O{r}"].value                  # O: 출자약정금액 (committed capital)
        rate = fee_ws[f"P{r}"].value                       # P: 수수료% (fee rate)
        annual = fee_ws[f"Q{r}"].value                     # Q: 연간 금액 (annual fee)
        if not isinstance(rate, (int, float)):             # skip 기타/non-numeric rows
            continue
        slug = _slug(str(name))
        rate_id = f"Metric:fund_fee_rate:{slug}"
        cap_id = f"Metric:fund_committed_capital:{slug}"
        fee_id = f"Metric:fund_mgmt_fee:{slug}"
        g.add_node(rate_id, type="Metric", label=f"{name} — fee rate", label_ko=str(name),
                   category="assumption", sheet="관리수수료", value=rate)
        g.add_node(cap_id, type="Metric", label=f"{name} — committed capital", label_ko=str(name),
                   category="driver", sheet="관리수수료", value=committed)
        g.add_node(fee_id, type="Metric", label=f"{name} — management fee", label_ko=str(name),
                   category="revenue", sheet="관리수수료", value=annual)
        for mid, col in ((rate_id, "P"), (cap_id, "O"), (fee_id, "Q")):
            cid = f"관리수수료!{col}{r}"
            g.add_node(cid, type="Cell", sheet="관리수수료")
            g.add_edge(mid, cid, type="DEFINED_IN")
        # committed capital × fee rate -> this fund's fee -> the aggregate management fee
        g.add_edge(cap_id, fee_id, type="DRIVES")
        g.add_edge(rate_id, fee_id, type="ASSUMPTION_OF")
        g.add_edge(fee_id, "Metric:management_fee", type="DRIVES")
        fund_node = FUND_NODE_MAP.get(str(name).strip())
        if fund_node and f"Fund:{fund_node}" in g:
            for mid in (rate_id, cap_id, fee_id):
                g.add_edge(mid, f"Fund:{fund_node}", type="BELONGS_TO")

    _attach_carry(g, wb)
    wb.close()

    # cross-metric edges, whitelist-guarded
    for src, dst in DRIVES:
        if src in METRIC_IDS and dst in METRIC_IDS:
            g.add_edge(f"Metric:{src}", f"Metric:{dst}", type="DRIVES")
    for src, dst in ASSUMPTION_OF:
        if src in METRIC_IDS and dst in METRIC_IDS:
            g.add_edge(f"Metric:{src}", f"Metric:{dst}", type="ASSUMPTION_OF")
    return g


def _attach_carry(g: nx.DiGraph, wb) -> None:
    """Per-fund GP carry & distribution from the `성과보수, 배당금` engine sheet.

    Mirrors the per-fund management-fee block: each fund gets a `fund_carry` metric (DTT on
    the node, MGT on ``value_mgt``) that DRIVES the aggregate ``performance_fee``, plus a
    `fund_distribution` metric and the three Exit assumptions (EBITDA, multiple, hurdle) that
    are ASSUMPTION_OF its carry. Every value is pinned to an exact cell via DEFINED_IN, and
    each metric BELONGS_TO its Biz Plan ``Fund:`` node.
    """
    ws = wb[CARRY_SHEET]

    def cell(ref: str) -> str:
        cid = f"{CARRY_SHEET}!{ref}"
        g.add_node(cid, type="Cell", sheet=CARRY_SHEET)
        return cid

    for f in CARRY_FUNDS:
        slug, v = _slug(f["alias"]), f["val"]
        fund_node = f"Fund:{f['node']}"

        carry_id = f"Metric:fund_carry:{slug}"
        g.add_node(carry_id, type="Metric", label=f"{f['alias']} — GP carry", label_ko="성과보수",
                   category="revenue", sheet=CARRY_SHEET, case="DTT",
                   value=ws[f"{v}6"].value, value_mgt=ws[f"{v}4"].value,
                   aliases=["성과보수", "carry", "carried interest", "performance fee", f["alias"]])
        g.add_edge(carry_id, cell(f"{v}6"), type="DEFINED_IN")     # DTT (active)
        g.add_edge(carry_id, cell(f"{v}4"), type="DEFINED_IN")     # MGT
        g.add_edge(carry_id, "Metric:performance_fee", type="DRIVES")

        dist_id = f"Metric:fund_distribution:{slug}"
        g.add_node(dist_id, type="Metric", label=f"{f['alias']} — distribution to GP",
                   label_ko="재산분배액", category="revenue", sheet=CARRY_SHEET, case="DTT",
                   value=ws[f"{v}9"].value, value_mgt=ws[f"{v}7"].value,
                   aliases=["재산분배액", "distribution", f["alias"]])
        g.add_edge(dist_id, cell(f"{v}9"), type="DEFINED_IN")
        g.add_edge(dist_id, cell(f"{v}7"), type="DEFINED_IN")

        # Exit assumptions (MGT-case row 5) that drive the carry computation.
        for suffix, ref, label_en, label_ko in (
            ("exit_ebitda", f["ebitda"], "Exit EBITDA/base", "Exit EBITDA"),
            ("exit_multiple", f["mult"], "Exit multiple", "Exit Multiple"),
            ("hurdle", f["hurdle"], "Hurdle rate", "기준수익률"),
        ):
            aid = f"Metric:fund_{suffix}:{slug}"
            g.add_node(aid, type="Metric", label=f"{f['alias']} — {label_en}", label_ko=label_ko,
                       category="assumption", sheet=CARRY_SHEET, value=ws[ref].value)
            g.add_edge(aid, cell(ref), type="DEFINED_IN")
            g.add_edge(aid, carry_id, type="ASSUMPTION_OF")
            if fund_node in g:
                g.add_edge(aid, fund_node, type="BELONGS_TO")

        if fund_node in g:
            g.add_edge(carry_id, fund_node, type="BELONGS_TO")
            g.add_edge(dist_id, fund_node, type="BELONGS_TO")


def _year_of(raw: object) -> int | str:
    """Pull a 4-digit year out of values like 2024, '2024E', '2024년 8월'."""
    import re
    m = re.search(r"(19|20)\d{2}", str(raw))
    return int(m.group(0)) if m else str(raw)


if __name__ == "__main__":
    from .. import FULL_WORKBOOK  # metric anchors live in the engine sheets, full model only

    g = nx.DiGraph()
    # standalone smoke test: build just the metric layer (no cell DAG needed)
    attach_metrics(g, FULL_WORKBOOK)
    metrics = [n for n, d in g.nodes(data=True) if d.get("type") == "Metric"]
    periods = sorted((d["year"] for n, d in g.nodes(data=True) if d.get("type") == "Period"),
                     key=lambda y: (1, 0) if isinstance(y, str) else (0, y))
    print(f"metrics: {len(metrics)}  periods: {periods}")
    print(f"edges: HAS_VALUE={sum(d['type']=='HAS_VALUE' for *_ ,d in g.edges(data=True))}, "
          f"DEFINED_IN={sum(d['type']=='DEFINED_IN' for *_ ,d in g.edges(data=True))}, "
          f"DRIVES={sum(d['type']=='DRIVES' for *_ ,d in g.edges(data=True))}, "
          f"ASSUMPTION_OF={sum(d['type']=='ASSUMPTION_OF' for *_ ,d in g.edges(data=True))}")
    print("\nscalar valuation/assumption metrics:")
    for n, d in g.nodes(data=True):
        if d.get("type") == "Metric" and d.get("category") in ("valuation", "assumption"):
            print(f"  {d['label']:32s} = {d.get('value')!r}  ({n.split(':',1)[1]})")
    print("\nEquity Value provenance chain (what DRIVES it):")
    ev = "Metric:equity_value"
    for u, v, d in g.in_edges(ev, data=True):
        print(f"  {u.split(':',1)[1]} --{d['type']}--> equity_value")
    print("\nmanagement_fee HAS_VALUE by period:")
    for u, v, d in g.out_edges("Metric:management_fee", data=True):
        if d["type"] == "HAS_VALUE":
            print(f"  {v.split(':',1)[1]}: {d['value']:.1f}  [{d['cell']}]")

    print("\nper-fund carry (DTT on node, MGT alt) -> performance_fee:")
    for n, d in g.nodes(data=True):
        if n.startswith("Metric:fund_carry:"):
            print(f"  {d['label']:24s} DTT={d.get('value')!r:>10}  MGT={d.get('value_mgt')!r}")
    pf_in = [u.split(":", 2)[-1] for u, v, d in g.in_edges("Metric:performance_fee", data=True)
             if d["type"] == "DRIVES" and u.startswith("Metric:fund_carry:")]
    print(f"  funds DRIVES performance_fee: {pf_in}")
