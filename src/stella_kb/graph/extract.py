"""Extract the cell-level formula dependency DAG from the workbook.

Each formula cell (e.g. ``='AUM Projection'!B12 + Tax!C4``) points at precedent cells.
Those precedent -> dependent links are the native ``DEPENDS_ON`` edges of the knowledge
graph. This module produces that DAG as a list of (precedent, dependent) cell pairs plus
per-cell metadata (formula string and cached value); graph.py lifts cells into semantic
nodes.

Cell ids are normalised to ``"Sheet!REF"`` (absolute ``$`` markers stripped, ranges
expanded to their corner cells — see expand_range).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter

# A reference inside a formula, optionally sheet-qualified:
#   B12   |   $B$12   |   Tax!C4   |   'AUM Projection'!B12   |   Sheet!A1:C9
_REF = re.compile(
    r"""(?:(?P<sheet>'[^']+'|[A-Za-z0-9_가-힣]+)!)?   # optional Sheet! (incl. Korean)
        (?P<a1>\$?[A-Z]{1,3}\$?\d+)                            # A1 or $A$1
        (?::(?P<a2>\$?[A-Z]{1,3}\$?\d+))?                      # optional :A2 (range)
    """,
    re.VERBOSE,
)


@dataclass
class CellInfo:
    """A single formula cell and the values it carries."""

    cell_id: str               # "Sheet!REF"
    formula: str               # raw formula string, leading "="
    value: object = None       # cached value (filled from a data_only pass), may be None
    precedents: list[str] = field(default_factory=list)


@dataclass
class DependencyGraph:
    cells: dict[str, CellInfo]          # cell_id -> CellInfo (formula cells only)
    edges: list[tuple[str, str]]        # (precedent_cell_id, dependent_cell_id)


def _norm(ref: str) -> str:
    return ref.replace("$", "")


def expand_range(sheet: str, a1: str, a2: str) -> list[str]:
    """Expand ``A1:C3`` into every ``Sheet!REF`` it covers."""
    min_c, min_r, max_c, max_r = range_boundaries(f"{_norm(a1)}:{_norm(a2)}")
    return [
        f"{sheet}!{get_column_letter(col)}{row}"
        for row in range(min_r, max_r + 1)
        for col in range(min_c, max_c + 1)
    ]


def parse_precedents(formula: str, current_sheet: str) -> list[str]:
    """Return the precedent cell ids referenced by ``formula``.

    References without a ``Sheet!`` qualifier resolve to ``current_sheet``. Ranges are
    expanded to individual cells so the DAG stays cell-grained.
    """
    out: list[str] = []
    for m in _REF.finditer(formula):
        sheet = m.group("sheet")
        sheet = sheet.strip("'") if sheet else current_sheet
        a1, a2 = m.group("a1"), m.group("a2")
        if a2:
            out.extend(expand_range(sheet, a1, a2))
        else:
            out.append(f"{sheet}!{_norm(a1)}")
    return out


def build_dependency_graph(path: str) -> DependencyGraph:
    """Read every formula cell in the workbook and assemble the dependency DAG.

    Two passes: ``data_only=False`` for formula strings (edges), ``data_only=True`` for
    Excel's cached results (node attributes). Cached values are ``None`` for cells never
    recalculated in Excel — recalc the file first if you need them populated.
    """
    wb_f = openpyxl.load_workbook(path, data_only=False, read_only=True)
    wb_v = openpyxl.load_workbook(path, data_only=True, read_only=True)

    cells: dict[str, CellInfo] = {}
    edges: list[tuple[str, str]] = []

    for ws in wb_f.worksheets:
        sheet = ws.title
        vws = wb_v[sheet]
        for row in ws.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.startswith("=")):
                    continue
                cell_id = f"{sheet}!{c.coordinate}"
                precs = parse_precedents(c.value, sheet)
                cells[cell_id] = CellInfo(
                    cell_id=cell_id,
                    formula=c.value,
                    value=vws[c.coordinate].value,
                    precedents=precs,
                )
                edges.extend((p, cell_id) for p in precs)

    wb_f.close()
    wb_v.close()
    return DependencyGraph(cells=cells, edges=edges)


if __name__ == "__main__":
    from .. import FULL_WORKBOOK  # graph layer needs the engine sheets absent from `_raw`

    dg = build_dependency_graph(FULL_WORKBOOK)
    print(f"formula cells: {len(dg.cells)}")
    print(f"dependency edges: {len(dg.edges)}")
    for cell_id, info in list(dg.cells.items())[:10]:
        print(f"  {cell_id}: {info.formula[:60]}  -> {len(info.precedents)} precedents")
