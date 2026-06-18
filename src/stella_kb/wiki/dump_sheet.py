"""Dump one or more sheets of the Project Stella workbook for analysis.

Usage (from repo root, with venv active):
    python -m src.stella_kb.wiki.dump_sheet "AUM Projection"
    python -m src.stella_kb.wiki.dump_sheet "DCF" --formulas      # show formula strings too
    python -m src.stella_kb.wiki.dump_sheet --list                # list all sheet names

Prints, for every non-empty cell, the coordinate, cached value, and (with
--formulas) the underlying formula. Designed to be piped/read by an analysis agent.
"""

import sys

import openpyxl

from .. import WORKBOOK


def list_sheets() -> None:
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    for i, name in enumerate(wb.sheetnames):
        print(f"{i:2d}  {name}")
    wb.close()


def dump(sheet: str, show_formulas: bool) -> None:
    wb_v = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    wb_f = (
        openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=False)
        if show_formulas
        else None
    )
    if sheet not in wb_v.sheetnames:
        print(f"!! sheet not found: {sheet!r}")
        print("   available:", ", ".join(wb_v.sheetnames))
        return
    vs = wb_v[sheet]
    fs = wb_f[sheet] if wb_f else None
    print(f"=== {sheet}  ({vs.max_row} rows x {vs.max_column} cols) ===")
    for row in vs.iter_rows():
        for c in row:
            if c.value is None:
                continue
            line = f"{c.coordinate}: {c.value!r}"
            if fs is not None:
                fv = fs[c.coordinate].value
                if isinstance(fv, str) and fv.startswith("="):
                    line += f"   [={fv[1:]}]"
            print(line)
    wb_v.close()
    if wb_f:
        wb_f.close()


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] == "--list":
        list_sheets()
    else:
        show_formulas = "--formulas" in args
        sheets = [a for a in args if not a.startswith("--")]
        for s in sheets:
            dump(s, show_formulas)
            print()
