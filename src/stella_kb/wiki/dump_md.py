"""Dump workbook sheets to Markdown, preserving the 2D grid + cell refs.

Unlike ``dump_sheet.py`` (a flat ``coord: value`` list), this emits a Markdown
table whose columns are the Excel column letters (A, B, C, ...) and whose rows
are the Excel row numbers — so the spatial layout an LLM needs to read structure
(which row is a header, which column is which fiscal year) is preserved, and any
cell ``Sheet!B7`` is reconstructable from the grid coordinates.

Each page has three parts:
  - **Values** — the cached-value grid (bounded to the used range).
  - **Merged ranges** — so spanned headers are explicit.
  - **Formulas** — every formula cell as ``REF = formula`` (the dependency detail
    kept out of the grid to keep it readable).

Numbers come from openpyxl; this dump never invents or transcribes values beyond
what the workbook holds. Cached values are ``None`` for cells Excel never
recalculated — recalc the .xlsx in LibreOffice first if the grid looks empty.

Usage (from repo root, with venv active):
    python -m src.stella_kb.wiki.dump_md --list                 # list sheet names
    python -m src.stella_kb.wiki.dump_md "DCF"                  # one sheet -> data/md/DCF.md
    python -m src.stella_kb.wiki.dump_md "DCF" "AUM Projection" # several
    python -m src.stella_kb.wiki.dump_md --all                  # every sheet
"""

from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from ..config import wiki_md_dir, wiki_workbook

WORKBOOK = wiki_workbook()
OUT_DIR = wiki_md_dir()


def _fmt(value: object) -> str:
    """Render a cell value compactly and table-safe (no stray pipes/newlines)."""
    if value is None:
        return ""
    if isinstance(value, datetime) and value.time() == datetime.min.time():
        text = value.date().isoformat()      # midnight -> bare date
    elif isinstance(value, (datetime, date)):
        text = value.isoformat()
    elif isinstance(value, float):
        text = f"{value:g}"          # 1000.0 -> "1000", 1234.56 -> "1234.56"
    else:
        text = str(value)
    return text.replace("|", "\\|").replace("\n", " ").strip()


def _used_bounds(ws) -> tuple[int, int, int, int]:
    """Tight (min_row, min_col, max_row, max_col) over cells with content."""
    min_r = min_c = None
    max_r = max_c = 0
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            r, col = c.row, c.column
            min_r = r if min_r is None else min(min_r, r)
            min_c = col if min_c is None else min(min_c, col)
            max_r, max_c = max(max_r, r), max(max_c, col)
    if min_r is None:
        return (1, 1, 0, 0)          # empty sheet
    return (min_r, min_c, max_r, max_c)


def sheet_to_md(name: str, vs, fs) -> str:
    """Build the Markdown page for one sheet (values grid + merged + formulas)."""
    min_r, min_c, max_r, max_c = _used_bounds(vs)
    lines = [f"# {name}", ""]

    if max_r == 0:
        lines.append("_(empty sheet)_")
        return "\n".join(lines) + "\n"

    lines.append(f"> {max_r - min_r + 1} rows x {max_c - min_c + 1} cols "
                 f"(used range {get_column_letter(min_c)}{min_r}:"
                 f"{get_column_letter(max_c)}{max_r})")
    lines += ["", "## Values", ""]

    cols = list(range(min_c, max_c + 1))
    header = "| | " + " | ".join(get_column_letter(c) for c in cols) + " |"
    sep = "|---|" + "---|" * len(cols)
    lines += [header, sep]
    for r in range(min_r, max_r + 1):
        cells = (_fmt(vs.cell(row=r, column=c).value) for c in cols)
        lines.append(f"| **{r}** | " + " | ".join(cells) + " |")

    merged = [str(rng) for rng in vs.merged_cells.ranges]
    if merged:
        lines += ["", "## Merged ranges", ""]
        lines += [f"- `{m}`" for m in sorted(merged)]

    formulas = []
    for row in fs.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                formulas.append((c.coordinate, v))
    if formulas:
        lines += ["", "## Formulas", ""]
        lines += [f"- `{coord}` = `{f}`" for coord, f in formulas]

    return "\n".join(lines) + "\n"


def main(argv: list[str]) -> None:
    if not argv or argv[0] == "--list":
        wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
        for i, n in enumerate(wb.sheetnames):
            print(f"{i:2d}  {n}")
        wb.close()
        return

    wb_v = openpyxl.load_workbook(WORKBOOK, data_only=True)
    wb_f = openpyxl.load_workbook(WORKBOOK, data_only=False)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if argv[0] == "--all":
        names = wb_v.sheetnames
    else:
        names = [a for a in argv if not a.startswith("--")]

    for name in names:
        if name not in wb_v.sheetnames:
            print(f"!! sheet not found: {name!r}")
            continue
        md = sheet_to_md(name, wb_v[name], wb_f[name])
        out = OUT_DIR / f"{name.replace('/', '_')}.md"
        out.write_text(md, encoding="utf-8")
        print(f"wrote {out}  ({len(md)} chars)")

    wb_v.close()
    wb_f.close()


if __name__ == "__main__":
    main(sys.argv[1:])
