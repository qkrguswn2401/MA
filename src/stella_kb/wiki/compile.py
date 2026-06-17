"""Compile pass: parsed structure + workbook + formula DAG -> one wiki page per sheet.

Stage 3 of the vectorless / LLM-wiki pipeline. Hybrid by design (same split as
``query.py``): **code builds everything it can prove, the LLM only writes prose.**

  - Facts table  — each grounded line item's value series, read straight from the
    workbook at ``value_row x axis-columns`` (openpyxl, never the model). Every number
    carries its ``Sheet!Ref``.
  - Cross-links  — the formula dependency DAG (``extract.py``) collapsed to the sheet
    level: ``[[depends on]]`` / ``[[feeds into]]``. Whitelist = the real sheet names, so
    no link can point at a page that doesn't exist (OpenKB pattern).
  - Aliases      — the KO/EN labels from the parse, lifted into the page header so the
    page is matchable by a bilingual query term (the words->node resolver).
  - Prose        — one LLM call writes the "What this is" paragraph, grounded in the
    facts table; it is told to cite cells and never invent a number. Skippable with
    ``--no-llm`` to inspect the deterministic scaffold first.

Input  : data/parsed/<sheet>.json  (from parse_llm.py)  +  WORKBOOK  +  the DAG
Output : data/wiki/pages/<sheet>.md

Usage (from repo root, venv active):
    python -m src.stella_kb.wiki.compile "DCF 장표 #2_DTT"     # one page
    python -m src.stella_kb.wiki.compile --all                 # every parsed sheet
    python -m src.stella_kb.wiki.compile --all --no-llm         # scaffold only, no LLM
"""

from __future__ import annotations

import concurrent.futures as cf
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from .. import config
from ..graph.extract import build_dependency_graph
from ..llm import chat
from ..prompts import load as load_prompt

from ..config import wiki_pages_dir, wiki_parsed_dir, wiki_workbook

WORKBOOK = wiki_workbook()
PARSED_DIR = wiki_parsed_dir()
OUT_DIR = wiki_pages_dir()


# --------------------------------------------------------------------------- helpers

def _fmt(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime) and value.time() == datetime.min.time():
        return value.date().isoformat()
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        if value == int(value):                       # 1.81e9 -> "1,812,940,000"
            return f"{int(value):,}"
        return f"{value:,.4f}".rstrip("0").rstrip(".")  # keep rates like 0.015
    return str(value).replace("|", "\\|").replace("\n", " ").strip()


# --- currency / unit -------------------------------------------------------------------

# The parsed unit string varies per sheet for the same meaning (원 / KRW / KRWm / KRWmn /
# 백만원 / Mixed / KRW/USD). Collapse the equivalents to one unambiguous token so every page
# states currency identically — but keep the won-vs-millions distinction (a 10^6 gap that
# would silently corrupt any valuation number read off the page).
_UNIT_GLOSS = {
    "KRW": "원 (KRW)",
    "KRWm": "백만원 (KRW millions)",
    "KRW·USD": "원·달러 혼용 (KRW & USD)",
    "Mixed": "혼합 단위 (mixed: %, 인구수, $ 등)",
}


def canon_unit(raw: object) -> str:
    """Collapse a parsed unit string to a canonical currency/scale token (or '' if absent)."""
    if not raw:
        return ""
    s = str(raw).strip()
    low = s.casefold().replace(" ", "")
    if "백만" in s or low in {"krwm", "krwmn", "krwmm", "krwmil", "krwmillion"}:
        return "KRWm"
    if "usd" in low and ("krw" in low or "원" in s):
        return "KRW·USD"
    if low in {"krw", "won", "원"}:
        return "KRW"
    if low == "mixed" or "혼합" in s:
        return "Mixed"
    return s  # unknown unit — pass through unchanged rather than guess


def unit_display(raw: object) -> str:
    """Human, bilingual currency/unit string for the page callout."""
    c = canon_unit(raw)
    if not c:
        return "단위 미상 (unit not stated in source)"
    return _UNIT_GLOSS.get(c, c)


def _local_ccy(sheet: str) -> str:
    """The local currency (LCU) for an EIU-style macro sheet."""
    if "(US)" in sheet:
        return "USD"
    if "(KR)" in sheet:
        return "KRW"
    return "LCU"


def row_unit(label: object, sheet: str) -> str:
    """Per-row currency/unit for a 'Mixed' macro sheet, read from the label's wording.

    EIU labels embed the unit: ``Real GDP (LCU)``, ``Nominal GDP (US$)``, ``Budget balance
    (% of GDP)``, ``Exchange rate LCU:US$``. LCU resolves to the sheet's local currency
    (KRW for the KR sheet, USD for the US sheet) so currency is never left implicit.
    """
    s = str(label or "")
    low = s.casefold()
    if "exchange rate" in low or "환율" in s or "lcu:us$" in low:
        return f"FX ({_local_ccy(sheet)}:US$)"
    if "%" in s:                                   # before US$: "% HHs … >US$100k" is a percentage
        return "%"
    if "us$" in low or "us dollar" in low or "달러" in s:
        return "USD"
    if "lcu" in low or "현지통화" in s:
        return _local_ccy(sheet)
    if "population" in low or "인구" in s:
        return "명 (persons)"
    return ""


def page_currency(meta: dict, items: list, sheet: str) -> tuple[str, str, dict]:
    """Resolve a page's currency presentation — never the bare word 'Mixed'.

    Returns ``(token, callout, per_row)``:
      - ``token``   — compact currency label for frontmatter + the index ToC.
      - ``callout`` — bilingual currency line shown under the page title.
      - ``per_row`` — ``{cell: unit}`` when the sheet genuinely mixes currencies (else
        ``{}``), which adds a per-row ``unit`` column to the Line-items table.
    """
    base = canon_unit(meta.get("unit"))
    if base != "Mixed":
        return base or "n/a", unit_display(meta.get("unit")), {}
    per_row = {it.get("label_cell"): row_unit(it.get("label"), sheet) for it in items
               if row_unit(it.get("label"), sheet)}
    seen = list(dict.fromkeys(per_row.values()))                 # distinct, in first-seen order
    ccys = [c for c in ("KRW", "USD") if c in seen]
    token = ("·".join(ccys) + " 등") if ccys else "행별 단위"
    callout = ("행별 단위 (per-row — 아래 표의 `unit` 열 참조) · 포함 단위: " + ", ".join(seen)
               if seen else "행별 단위 (per-row — 아래 표의 `unit` 열 참조)")
    return token, callout, per_row


def load_values(sheet: str) -> dict[str, object]:
    """``{coordinate: cached value}`` for one sheet (ground truth for the facts table)."""
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb[sheet]
    vals = {c.coordinate: c.value for row in ws.iter_rows() for c in row
            if c.value is not None}
    wb.close()
    return vals


def sheet_links() -> dict[str, dict[str, list[str]]]:
    """Collapse the cell-level DAG to sheet-level depends-on / feeds-into sets.

    Returns ``{sheet: {"depends_on": [...], "feeds_into": [...]}}`` for every sheet.
    """
    dg = build_dependency_graph(WORKBOOK)
    links: dict[str, dict[str, set]] = {}
    for prec, dep in dg.edges:
        sp, sd = prec.rsplit("!", 1)[0], dep.rsplit("!", 1)[0]
        if sp == sd:
            continue
        links.setdefault(sd, {"depends_on": set(), "feeds_into": set()})
        links.setdefault(sp, {"depends_on": set(), "feeds_into": set()})
        links[sd]["depends_on"].add(sp)
        links[sp]["feeds_into"].add(sd)
    return {s: {k: sorted(v) for k, v in d.items()} for s, d in links.items()}


def tables_of(parsed: dict) -> list[dict]:
    """Normalize a parsed sheet to a list of ``{title, year_axis, line_items}`` tables.

    A sheet may stack several tables on different axes (a monthly roster + an annual 월평균
    table); the parser emits those under ``tables``. A plain single-axis sheet is returned
    as one untitled table, so every consumer can iterate uniformly.
    """
    if parsed.get("tables"):
        return [{"title": t.get("title"), "year_axis": t.get("year_axis") or {},
                 "line_items": t.get("line_items") or []} for t in parsed["tables"]]
    return [{"title": None, "year_axis": parsed.get("year_axis") or {},
             "line_items": parsed.get("line_items") or []}]


def usable_tables(parsed: dict, vals: dict) -> list[dict]:
    """Tables worth rendering: those that resolve to a real time series (≥2 distinct periods).

    A multi-table sheet can yield a degenerate "table" — a qualitative block (e.g. an org
    chart listing job titles per year, with no clean numeric axis) collapses to <2 periods.
    Dropping those keeps such junk off the page (and out of the alias index). Never returns
    empty: a legitimate single-table sheet with a thin axis is kept as-is.
    """
    tables = tables_of(parsed)
    keep = [t for t in tables if len(periods_of(t.get("year_axis") or {}, vals)[1]) >= 2]
    return keep or tables


def all_items(parsed: dict, vals: dict | None = None) -> list[dict]:
    """Every line item across a sheet's tables (for aliases / currency / index).

    With ``vals`` given, restricts to :func:`usable_tables` so dropped degenerate tables
    don't leak their labels into the alias index; without it, spans every table.
    """
    tables = usable_tables(parsed, vals) if vals is not None else tables_of(parsed)
    return [it for t in tables for it in t["line_items"]]


def periods_of(axis: dict, vals: dict) -> tuple[dict, list[str], list[str]]:
    """``(axis_cols, periods, period_labels)`` for one table's axis.

    Periods are deduped on the resolved period, keeping the **last column** per period. For a
    monthly date axis this is the period-end collapse (Dec, or the last available month) —
    turning a roster's monthly Total row into a clean year-end series. For the usual one
    column per year it is a no-op. ``period_labels`` show the kept column's own header text
    (``Dec-20``, ``FY23``, the basis date) so month/half granularity stays visible.
    """
    axis_cols = axis.get("columns") or {}
    axis_row = axis.get("row")
    last_col = {}
    for col, period in axis_cols.items():
        last_col[str(period)] = col            # later columns overwrite -> period-end wins
    periods = list(last_col)
    labels = []
    for p in periods:
        hdr = vals.get(f"{last_col[p]}{axis_row}") if axis_row else None
        labels.append(_fmt(hdr) if hdr not in (None, "") else p)
    return axis_cols, periods, labels


def value_series(vals: dict, item: dict, axis_cols: dict) -> list[tuple[str, object, str]]:
    """``(period_label, value, cell_ref)`` for a line item, read from real cells.

    Prefers the year axis; if the item's row has no values under the axis columns
    (e.g. a scalar summary line), falls back to the non-empty cells to the right of
    the label cell.
    """
    row = item.get("value_row")
    if not row:
        return []
    out = []
    for col, year in (axis_cols or {}).items():
        v = vals.get(f"{col}{row}")
        if v is not None:
            out.append((str(year), v, f"{col}{row}"))
    if out:
        return out
    label_col = column_index_from_string(item["label_cell"].rstrip("0123456789"))
    for coord, v in vals.items():
        c, r = coord.rstrip("0123456789"), coord[len(coord.rstrip("0123456789")):]
        if r == str(row) and column_index_from_string(c) > label_col \
                and isinstance(v, (int, float)) and not isinstance(v, bool):
            out.append((c, v, coord))
    out.sort(key=lambda t: column_index_from_string(t[0]) if t[0].isalpha() else 0)
    return out


# --------------------------------------------------------------------------- prose

_PROSE_SYS = load_prompt("wiki_prose_system")


def _context(sheet: str) -> str:
    """A one-line 분류 gloss (section/group/kind/case) to anchor the prose.

    Lets the LLM write a *discriminating* lead sentence — and describe a sheet from its
    identity when its facts are thin — instead of a generic "이 시트는 ~를 나타냅니다".
    Deferred import: ``index`` imports from ``compile`` at module load, so importing
    ``classify`` at top level here would form a cycle.
    """
    from .index import classify  # deferred to avoid the index<->compile import cycle

    c = classify(sheet)
    gloss = " / ".join(p for p in (c.get("section"), c.get("group"), c.get("kind")) if p)
    line = f"분류: {gloss}" if gloss else ""
    if c.get("case"):
        line += f"  (케이스: {c['case']})"
    return line


def _prose(sheet: str, meta: dict, facts: list[str]) -> str:
    body = "\n".join(facts) or "(사실 항목 없음 — 분류 정보만으로 이 표의 성격·용도를 설명할 것)"
    ctx = _context(sheet)
    user = (f"Sheet: {sheet!r}  (case={meta.get('case')}, unit={meta.get('unit')})\n"
            f"{ctx}\n"
            f"Facts:\n{body}\n\nSummary:")
    try:
        return chat([{"role": "system", "content": _PROSE_SYS},
                     {"role": "user", "content": user}], max_tokens=300).strip()
    except Exception as e:  # noqa: BLE001 — prose is best-effort; scaffold still stands
        return f"_(prose unavailable: {type(e).__name__})_"


# --------------------------------------------------------------------------- compile

def _render_table(table: dict, vals: dict, per_row: dict) -> tuple[list[str], list[str]]:
    """Build one table's facts rows (markdown) + a compact facts block for the prose call."""
    axis_cols, periods, period_labels = periods_of(table.get("year_axis") or {}, vals)
    table_rows, facts_lines = [], []
    for it in table.get("line_items") or []:
        series = value_series(vals, it, axis_cols)
        # each period carries its own value cell ([J6]) so the agent can cite the exact period
        # cell, not just the row's label cell — without this the retriever can only pull one
        # scalar per row (the cell-on-page guard rejects any per-period cell it can't see).
        sval = {p: f"{_fmt(v)} [{ref}]" for p, v, ref in series}
        if periods and any(p in sval for p in periods):
            cells = " | ".join(sval.get(p, "") for p in periods)
        elif periods:
            # off-axis scalar (e.g. the DCF summary box: EV/Equity in column E, not on the
            # year axis). value_series found it but its key isn't a year — surface the
            # value(s) with the cell ref in the first column rather than blanking the row.
            joined = ", ".join(f"{_fmt(v)} [{ref}]" for _, v, ref in series)
            cells = " | ".join([joined] + [""] * (len(periods) - 1))
        else:
            cells = " | ".join(f"{p}={_fmt(v)} [{ref}]" for p, v, ref in series)
        ucol = f" {per_row.get(it.get('label_cell'), '')} |" if per_row else ""
        table_rows.append(
            f"| {it.get('label','')} | {it.get('label_ko') or ''} | "
            f"{it.get('role','')} |{ucol} `{it.get('label_cell','')}` | {cells} |")
        if series:
            facts_lines.append(
                f"- {it.get('label','')} ({it.get('label_ko') or ''}) "
                f"[{it.get('label_cell')}]: "
                + ", ".join(f"{p}={_fmt(v)} [{ref}]" for p, v, ref in series))

    ucol_h = " unit |" if per_row else ""
    ucol_s = "---|" if per_row else ""
    header = f"| Item | KO | role |{ucol_h} cell |" + ("".join(f" {p} |" for p in period_labels)
                                                       if period_labels else " values |")
    sep = f"|---|---|---|{ucol_s}---|" + ("---|" * len(periods) if periods else "---|")
    return [header, sep, *table_rows], facts_lines


def compile_page(sheet: str, parsed: dict, vals: dict,
                 links: dict, whitelist: set, use_llm: bool) -> str:
    meta = parsed.get("meta") or {}
    tables = usable_tables(parsed, vals)
    items = all_items(parsed, vals)

    # gather aliases for the header (the words->node resolver)
    aliases = []
    for it in items:
        for a in [it.get("label_ko"), it.get("label_en"), *(it.get("aliases") or [])]:
            if a and a not in aliases:
                aliases.append(a)

    # currency/unit: a precise token + callout, and per-row units when the sheet mixes them
    ccy_token, ccy_callout, per_row = page_currency(meta, items, sheet)

    # one facts section per table (a sheet may stack several on different axes)
    sections, facts_lines = [], []
    for t in tables:
        rows, facts = _render_table(t, vals, per_row)
        title = t.get("title")
        heading = (f"## {title} — 단위/unit: {ccy_token}" if title
                   else f"## Line items — 단위/unit: {ccy_token}")
        sections += [heading + ("  (행별 단위는 아래 unit 열 참조)" if per_row else ""), "", *rows, ""]
        facts_lines += facts

    # --- assemble markdown ---
    out = ["---", f"sheet: {sheet}"]
    out.append(f"section: {meta.get('title') or ''}")
    if meta.get("case"):
        out.append(f"case: {meta['case']}")
    out.append(f"unit: {ccy_token}")
    if aliases:
        out.append("aliases: [" + ", ".join(aliases) + "]")
    out += ["---", "", f"# {meta.get('title') or sheet}", "",
            f"> **통화·단위 (Currency / Unit): {ccy_callout}**"
            + ("" if per_row else " — 이 페이지의 모든 금액 수치에 적용됩니다."), ""]

    out += ["## What this is", ""]
    out.append(_prose(sheet, meta, facts_lines) if use_llm
               else "_(scaffold only — run without --no-llm for prose)_")
    out.append("")

    out += sections

    link = links.get(sheet, {})
    dep = [f"[[{s}]]" for s in link.get("depends_on", []) if s in whitelist]
    feed = [f"[[{s}]]" for s in link.get("feeds_into", []) if s in whitelist]
    out += ["", "## Links", ""]
    out.append(f"- **Depends on:** {', '.join(dep) if dep else '—'}")
    out.append(f"- **Feeds into:** {', '.join(feed) if feed else '—'}")

    return "\n".join(out) + "\n"


if __name__ == "__main__":
    import sys

    args = sys.argv[1:]
    use_llm = "--no-llm" not in args
    names_arg = [a for a in args if not a.startswith("--")]

    parsed_files = {p.stem: p for p in PARSED_DIR.glob("*.json")}
    if "--all" in args:
        names = sorted(parsed_files)
    else:
        names = names_arg or ["DCF 장표 #2_DTT"]

    print("building sheet-level link graph from the formula DAG ...")
    links = sheet_links()
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    whitelist = set(wb.sheetnames)
    wb.close()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    def _compile_and_write(name: str) -> str:
        if name not in parsed_files:
            return f"!! no parsed JSON for {name!r} — run parse_llm first"
        parsed = json.loads(parsed_files[name].read_text(encoding="utf-8"))
        sheet = parsed.get("sheet", name)
        if sheet not in whitelist:  # _raw-only: skip any full-workbook-only sheet (e.g. carry)
            return f"-- skipping {name!r} — sheet {sheet!r} not in _raw (out of wiki scope)"
        md = compile_page(sheet, parsed, load_values(sheet), links, whitelist, use_llm)
        (OUT_DIR / f"{name}.md").write_text(md, encoding="utf-8")
        return (f"wrote {name}.md  ({len(parsed.get('line_items', []))} items, "
                f"{'prose' if use_llm else 'scaffold'})")

    # Prose calls run concurrently (bounded); scaffold-only is CPU-light but harmless.
    workers = (1 if not use_llm
               else max(1, min(config.parse_concurrency(), len(names))))
    print(f"compiling {len(names)} pages with {workers} worker(s) ...")
    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        for msg in ex.map(_compile_and_write, names):
            print(msg, flush=True)
