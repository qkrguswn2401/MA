"""Deterministic transaction-ledger extraction + query — for ``*_거래내역`` sheets.

These sheets are **transaction ledgers**, not time-series tables: hundreds of rows, each a
single transaction (일시·거래처·적요·입금·출금·잔고), often with **two side-by-side sub-ledgers**
(a KRW band and a USD band, the USD band carrying 적용환율·원화). The LLM parse pass (built for
fiscal-year tables) finds no year axis and drops every row — so the wiki page ends up with only
column headers and the agent can't answer "find/sum transactions matching <적요 keyword>".

This module bypasses the LLM entirely: it reads the raw rows with openpyxl into structured
``TransactionRow`` records, and ``query_ledger`` filters by 적요 keyword and **sums 출금 per
currency deterministically** (no LLM arithmetic) with per-row cell provenance — the auditable
answer to questions like Q19 ("그 펀드 거래내역에서 이 금액을 찾을 수 있나").

Build → ``write_ledgers`` emits one ``<sheet>.json`` sidecar per 거래내역 sheet under the wiki's
``ledgers/`` dir; the agent's ``query_ledger`` tool reads those at query time.
"""
from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

# header tokens (substring match, tolerant of trailing markers/newlines)
_DESC = "적요"
_OUT = "출금"
_IN = "입금"
_BAL = "잔고"
_DATE = "일시"
_PARTY = "거래처"
_KRWEQ = "원화"
_HEADER_SCAN_ROWS = 6  # headers sit in the first few rows (row 1 (KRW)/(USD), rows 2-3 labels)


@dataclass
class TransactionRow:
    """One ledger transaction. ``ref`` is the 출금(or 입금) cell — the citable provenance."""

    fund: str
    currency: str          # "KRW" | "USD"
    row: int
    date: str | None
    party: str | None      # 거래처
    desc: str | None       # 적요
    inflow: float | None   # 입금
    outflow: float | None  # 출금
    balance: float | None  # 잔고
    krw_equiv: float | None  # 원화 (USD band only)
    ref: str               # e.g. "차이나1호_거래내역!E15" (the 출금 cell)


def _find_cols(ws, sheet: str) -> list[dict]:
    """Detect each currency sub-ledger's column band from the header rows.

    Returns ``[{currency, date, party, desc, inflow, outflow, balance, krw_equiv, c0, c1,
    header_row}]`` — one per band. Bands are split at each 적요 column; currency is USD if the
    band carries a 원화 column, else KRW (the (KRW)/(USD) row-1 markers corroborate)."""
    maxc = ws.max_column
    # column -> set of header tokens seen in the scan rows
    htok: dict[int, str] = {}
    krw_marker_col = usd_marker_col = None
    for r in range(1, _HEADER_SCAN_ROWS + 1):
        for c in range(1, maxc + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).replace("\n", "").strip()
            if "(KRW)" in s.upper() or s.upper() == "KRW":
                krw_marker_col = c
            if "(USD)" in s.upper() or s.upper() == "USD":
                usd_marker_col = c
            for tok in (_DESC, _OUT, _IN, _BAL, _DATE, _PARTY, _KRWEQ):
                if tok in s:
                    htok.setdefault(c, tok)  # first header token wins per column

    desc_cols = sorted(c for c, t in htok.items() if t == _DESC)
    if not desc_cols:
        return []
    bands: list[dict] = []
    out_header_row = 1
    for i, dc in enumerate(desc_cols):
        c1 = desc_cols[i + 1] - 1 if i + 1 < len(desc_cols) else maxc
        # gather this band's columns (from a bit before 적요 to the next 적요)
        in_band = lambda tok, lo=max(1, dc - 3), hi=c1: next(  # noqa: E731
            (c for c, t in htok.items() if t == tok and lo <= c <= hi), None)
        oc = in_band(_OUT)
        # find the header row of 출금/입금 (the data starts just below it)
        if oc:
            for r in range(1, _HEADER_SCAN_ROWS + 1):
                cv = ws.cell(r, oc).value
                if cv and _OUT in str(cv):
                    out_header_row = max(out_header_row, r)
                    break
        krweq = in_band(_KRWEQ)
        currency = "USD" if krweq or (usd_marker_col and dc >= usd_marker_col
                                      and (krw_marker_col is None or usd_marker_col > krw_marker_col)) else "KRW"
        bands.append({
            "currency": currency, "desc": dc, "outflow": oc, "inflow": in_band(_IN),
            "balance": in_band(_BAL), "date": in_band(_DATE), "party": in_band(_PARTY),
            "krw_equiv": krweq, "c1": c1,
        })
    # data starts after the deepest header row
    for b in bands:
        b["header_row"] = out_header_row
    return bands


def _num(v: Any) -> float | None:
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def extract_ledger(ws, sheet: str, fund: str) -> list[TransactionRow]:
    """Read every transaction row of a 거래내역 worksheet into ``TransactionRow`` records."""
    bands = _find_cols(ws, sheet)
    rows: list[TransactionRow] = []
    for b in bands:
        if not b["desc"]:
            continue
        for r in range(b["header_row"] + 1, ws.max_row + 1):
            desc = ws.cell(r, b["desc"]).value if b["desc"] else None
            inflow = _num(ws.cell(r, b["inflow"]).value) if b["inflow"] else None
            outflow = _num(ws.cell(r, b["outflow"]).value) if b["outflow"] else None
            if desc is None and inflow is None and outflow is None:
                continue  # blank row in this band
            ref_col = b["outflow"] or b["inflow"] or b["desc"]
            rows.append(TransactionRow(
                fund=fund, currency=b["currency"], row=r,
                date=str(ws.cell(r, b["date"]).value) if b["date"] else None,
                party=(str(ws.cell(r, b["party"]).value) if b["party"]
                       and ws.cell(r, b["party"]).value is not None else None),
                desc=str(desc).strip() if desc is not None else None,
                inflow=inflow, outflow=outflow,
                balance=_num(ws.cell(r, b["balance"]).value) if b["balance"] else None,
                krw_equiv=_num(ws.cell(r, b["krw_equiv"]).value) if b["krw_equiv"] else None,
                ref=f"{sheet}!{get_column_letter(ref_col)}{r}",
            ))
    return rows


def query_ledger_rows(rows: list[dict | TransactionRow], keywords: list[str],
                      field: str = "outflow") -> dict:
    """Filter rows whose 적요 contains any keyword; sum ``field`` per currency. Deterministic.

    Returns ``{matched: [...], totals: {KRW, USD}, totals_krw: {...}, absent: [kw,...],
    keywords}``. ``totals`` sums the raw ``field`` per currency; ``totals_krw`` converts the USD
    band via each row's 원화(krw_equiv) so KRW and USD are comparable in won. ``absent`` lists the
    keywords with **zero** matching rows — the "그런 적요 자체가 없다" signal Q19 needs."""
    rs = [asdict(r) if isinstance(r, TransactionRow) else dict(r) for r in rows]
    kws = [k for k in keywords if k]
    matched: list[dict] = []
    totals: dict[str, float] = {}
    totals_krw: dict[str, float] = {}
    hit_per_kw: dict[str, int] = {k: 0 for k in kws}
    for r in rs:
        desc = r.get("desc") or ""
        which = [k for k in kws if k in desc]
        if not which:
            continue
        for k in which:
            hit_per_kw[k] += 1
        amt = r.get(field)
        cur = r.get("currency", "KRW")
        if isinstance(amt, (int, float)):
            totals[cur] = totals.get(cur, 0.0) + amt
            krw = r.get("krw_equiv") if cur == "USD" and r.get("krw_equiv") else (amt if cur == "KRW" else None)
            if isinstance(krw, (int, float)):
                totals_krw[cur] = totals_krw.get(cur, 0.0) + krw
        matched.append({"currency": cur, "desc": desc, field: amt,
                        "krw_equiv": r.get("krw_equiv"), "ref": r.get("ref"),
                        "date": r.get("date"), "party": r.get("party")})
    return {
        "keywords": kws,
        "matched": matched,
        "totals": {c: round(v) for c, v in totals.items()},
        "totals_krw": {c: round(v) for c, v in totals_krw.items()},
        "krw_total": round(sum(totals_krw.values())) if totals_krw else 0,
        "absent": [k for k in kws if hit_per_kw[k] == 0],
    }


# --- build-time sidecar emission -----------------------------------------------------------

def write_ledgers(workbook: str, sheets: list[str], out_dir: Path) -> dict[str, int]:
    """Extract every ``*_거래내역`` sheet in ``sheets`` to ``out_dir/<sheet>.json``. Returns
    ``{sheet: row_count}``. Fund = sheet name minus the ``_거래내역`` suffix."""
    import openpyxl

    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    counts: dict[str, int] = {}
    try:
        for sheet in sheets:
            if not sheet.endswith("_거래내역") or sheet not in wb.sheetnames:
                continue
            fund = sheet[: -len("_거래내역")]
            # read_only worksheets need materializing for random cell access
            ws = wb[sheet]
            rows = extract_ledger(_Materialized(ws), sheet, fund)
            (out_dir / f"{sheet}.json").write_text(
                json.dumps([asdict(r) for r in rows], ensure_ascii=False), encoding="utf-8")
            counts[sheet] = len(rows)
    finally:
        wb.close()
    return counts


class _Materialized:
    """openpyxl read_only ws → random-access shim (read_only cells aren't indexable)."""

    def __init__(self, ws: Any) -> None:
        self._rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        self.max_row = len(self._rows)
        self.max_column = max((len(r) for r in self._rows), default=0)

    def cell(self, row: int, column: int):
        class _C:
            __slots__ = ("value",)

            def __init__(self, v):
                self.value = v
        if 1 <= row <= self.max_row:
            r = self._rows[row - 1]
            if 1 <= column <= len(r):
                return _C(r[column - 1])
        return _C(None)
