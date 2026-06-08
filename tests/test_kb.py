"""KB-build helpers: sheet classification, the carry formatter, and the carry anchors.

The carry-value tests pin the curated cells against the workbook so a model revision that
shifts a column or row trips a test instead of silently corrupting an answer.
"""

from __future__ import annotations

import openpyxl
import pytest

from src.stella_kb.wiki.carry import FUNDS, SHEET, _carry_cells, _fmt
from src.stella_kb.wiki.index import classify


# --- classify: sheet name → section/group via tokens ----------------------------------


@pytest.mark.parametrize("name, section_contains", [
    ("성과보수, 배당금", "Fin.Model"),
    ("EIU(KR)", "Macro"),
    ("DCF 장표 #1_MGT", "Exhibits"),
    ("제8호_비용", "Biz Plan"),
    ("4.1BS", "BSPL"),
])
def test_classify_section(name, section_contains):
    assert section_contains in classify(name)["section"]


def test_classify_case_from_suffix():
    assert classify("DCF 장표 #1_MGT")["case"] == "MGT"
    assert classify("DCF 장표 #2_DTT")["case"] == "DTT"


# --- _fmt: faithful display of cell values --------------------------------------------


@pytest.mark.parametrize("value, shown", [
    (20048, "20,048"),
    (391912.1764, "391,912.1764"),     # carry: full precision, no truncation
    (13.744, "13.744"),                # exit multiple: precision preserved
    (0, "0"),
    (None, ""),
    ("2026-12-31 00:00:00", "2026-12-31"),
])
def test_fmt(value, shown):
    assert _fmt(value) == shown


# --- carry anchors map to the right value cells (FUNDS is the curated table) -----------


def test_carry_cells_layout():
    cells = _carry_cells(FUNDS[0])              # 제2호, value column E
    assert cells == {"carry_mgt": "E4", "carry_dtt": "E6",
                     "dist_mgt": "E7", "dist_dtt": "E9"}


def test_six_fund_blocks():
    assert [f["alias"] for f in FUNDS] == ["제2호", "옐로씨", "제5호", "제7호", "제7-1호", "제8호"]


def test_carry_values_against_workbook(full_workbook):
    """The two funds that clear the hurdle, pinned to their cells (guards model drift)."""
    ws = openpyxl.load_workbook(full_workbook, data_only=True, read_only=True)[SHEET]
    f7 = next(f for f in FUNDS if f["alias"] == "제7호")
    f8 = next(f for f in FUNDS if f["alias"] == "제8호")
    assert round(ws[_carry_cells(f7)["carry_mgt"]].value) == 391912   # AU4
    assert round(ws[_carry_cells(f7)["carry_dtt"]].value) == 135635   # AU6
    assert ws[_carry_cells(f8)["carry_mgt"]].value == 20048           # BW4
    # the four below-hurdle funds carry 0
    for alias in ("제2호", "옐로씨", "제5호", "제7-1호"):
        f = next(x for x in FUNDS if x["alias"] == alias)
        assert ws[_carry_cells(f)["carry_mgt"]].value == 0
