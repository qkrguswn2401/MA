"""Evaluate the **wiki-based index agent** on the stella_case PDF×Excel cross-check set.

The cross-check dataset (``test_data/rag_test_dataset/stella_case/``) ships its own
corpus — an 18-sheet cleaned Excel (``RAG_0604_테스트용_Input.xlsx``) plus a summary PDF —
distinct from the project's valuation-model wiki. This harness builds a **fresh, Excel-only**
wiki index over that test workbook (PDF intentionally excluded — see the README's 3-tier
design; T2/T3 cross-check items are expected to be hard without the PDF), runs the 20
questions through the wiki agent against it, and LLM-judges each answer tier-aware.

Everything is redirected to ``data/eval_stella/`` so the canonical ``data/wiki/`` (the
valuation-model wiki) is never touched.

    python -m eval.stella_crosscheck build   # 4 wiki stages over the test Excel
    python -m eval.stella_crosscheck eval    # run 20 questions -> answers json
    python -m eval.stella_crosscheck judge   # tier-aware LLM judge -> scoreboard
    python -m eval.stella_crosscheck all     # build -> eval -> judge
"""

from __future__ import annotations

import json
import os
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from src.stella_kb import config

ROOT = Path(__file__).resolve().parents[1]
# The golden cross-check case (Excel + FDD PDF + 20 ground-truth questions). The dataset is
# versioned under test_data/<ver>/; override CASE/questions/output via env so the same harness
# can score a different build (e.g. the v0.2 wiki) without code edits.
CASE = Path(os.environ.get(
    "EVAL_CASE", str(ROOT / "test_data" / "v0.1" / "rag_test_dataset" / "stella_case")))
TEST_XLSX = str(CASE / "files" / "RAG_0604_테스트용_Input.xlsx")
TEST_PDF = str(CASE / "files" / "Stella_FDD_ExecSummary_p4-17.pdf")
QUESTIONS = Path(os.environ.get("EVAL_QUESTIONS",
                                str(CASE / "ground_truth" / "cross_check_questions.jsonl")))

# Output dir (answers/scores/report). Separate from the *target wiki* so we can evaluate a
# prebuilt wiki (EVAL_WIKI) and write results elsewhere.
EVAL_DIR = Path(os.environ.get("EVAL_OUT_DIR", str(ROOT / "data" / "eval" / "stella_crosscheck")))
MD_DIR = EVAL_DIR / "md"
PARSED_DIR = EVAL_DIR / "parsed"
# The wiki the agent is evaluated against. Default: the self-built Excel-only wiki under
# EVAL_DIR; set EVAL_WIKI to score an already-built wiki (Excel+PDF), skipping `build`.
WIKI_DIR = Path(os.environ.get("EVAL_WIKI", str(EVAL_DIR / "wiki")))
PAGES_DIR = WIKI_DIR / "pages"
INDEX_JSON = WIKI_DIR / "index.json"
INDEX_MD = WIKI_DIR / "INDEX.md"

ANSWERS_JSON = EVAL_DIR / "answers.json"
SCORES_JSON = EVAL_DIR / "scores.json"
REPORT_MD = EVAL_DIR / "report.md"

# Which agent backend to exercise. "wiki" (default) = the wiki backend directly (the eval
# path so far). "auto" = the supervisor StateGraph (route → wiki/dart → compose/passthrough) —
# the path the UI actually uses, which the wiki path leaves UNMEASURED. For this Centroid set
# the supervisor should route every question to wiki and pass its answer through verbatim, so
# `auto` ≈ `wiki` is the pass condition; a gap is the regression signal (mis-route or compose
# erosion). Compare means over runs (the eval is noisy); hold the built pages fixed.
EVAL_SOURCE = os.environ.get("EVAL_SOURCE", "wiki")


# --- stage 1: build the Excel-only wiki over the test workbook --------------------------

def build() -> None:
    """Run dump_md -> parse_llm -> compile -> index against the test Excel, into EVAL_DIR.

    Each stage reads its module-global ``WORKBOOK`` and output-dir constants at call time,
    so we redirect by rebinding those globals — no edits to the pipeline modules.
    """
    import openpyxl

    from src.stella_kb.wiki import compile as wc
    from src.stella_kb.wiki import dump_md, index, parse_llm

    for d in (MD_DIR, PARSED_DIR, PAGES_DIR):
        d.mkdir(parents=True, exist_ok=True)

    # stage 1 — grids
    dump_md.WORKBOOK = TEST_XLSX
    dump_md.OUT_DIR = MD_DIR
    print("[1/4] dump_md --all")
    dump_md.main(["--all"])

    # stage 2 — LLM parse (grounded structural schema), concurrent + bounded
    parse_llm.WORKBOOK = TEST_XLSX
    parse_llm.MD_DIR = MD_DIR
    parse_llm.OUT_DIR = PARSED_DIR
    sheets = [p.stem for p in sorted(MD_DIR.glob("*.md")) if p.stat().st_size > 200]
    print(f"[2/4] parse_llm: {len(sheets)} sheets")
    with ThreadPoolExecutor(max_workers=6) as ex:
        for msg in ex.map(parse_llm._parse_and_write, sheets):
            print("   ", msg)

    # stage 3 — compile one wiki page per parsed sheet (deterministic facts + LLM prose)
    wc.WORKBOOK = TEST_XLSX
    wc.PARSED_DIR = PARSED_DIR
    wc.OUT_DIR = PAGES_DIR
    wb = openpyxl.load_workbook(TEST_XLSX, read_only=True)
    whitelist = set(wb.sheetnames)
    wb.close()
    links = wc.sheet_links()  # empty for this no-formula workbook, but harmless
    parsed_files = {p.stem: p for p in PARSED_DIR.glob("*.json")}
    print(f"[3/4] compile: {len(parsed_files)} pages")

    def _compile(name: str) -> str:
        parsed = json.loads(parsed_files[name].read_text(encoding="utf-8"))
        sheet = parsed.get("sheet", name)
        if sheet not in whitelist:
            return f"-- skip {name!r}: {sheet!r} not in workbook"
        md = wc.compile_page(sheet, parsed, wc.load_values(sheet), links, whitelist, True)
        (PAGES_DIR / f"{name}.md").write_text(md, encoding="utf-8")
        return f"wrote {name}.md"

    with ThreadPoolExecutor(max_workers=6) as ex:
        for msg in ex.map(_compile, sorted(parsed_files)):
            print("   ", msg)

    # stage 4 — index / ToC
    index.WORKBOOK = TEST_XLSX
    index.PARSED_DIR = PARSED_DIR
    index.PAGES_DIR = PAGES_DIR
    index.OUT_JSON = INDEX_JSON
    index.OUT_MD = INDEX_MD
    print("[4/4] index")
    idx = index.build_index()
    INDEX_JSON.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding="utf-8")
    INDEX_MD.write_text(index.render_md(idx), encoding="utf-8")
    print(f"   -> {INDEX_JSON}\n   -> {INDEX_MD}")
    print(f"   pages={len(idx['pages'])}  aliases={len(idx['alias_index'])}")

    from src.stella_kb.wiki.ledger import write_ledgers  # 거래내역 row sidecars
    led = write_ledgers(TEST_XLSX, [s for s in idx["pages"] if s.endswith("_거래내역")],
                        WIKI_DIR / "ledgers")
    print(f"   ledgers: {sum(led.values())} rows / {len(led)} sheet(s)")


def buildall() -> None:
    """Full rebuild with the Excel and PDF pipelines **overlapped**. They share nothing until
    the final index merge — Excel does parse→compile→index (writes index.json); PDF does
    convert→structure (writes FDD pages, returns index pieces) — so they run concurrently and
    the PDF pieces are merged into the freshly built index at the end. Wall-time = max(excel,
    pdf) instead of their sum, and merging PDF last also avoids the recompile-drops-PDF bug."""
    from src.stella_kb.wiki import index as wiki_index
    from src.stella_kb.wiki import pdf_pages

    pieces: dict = {}

    def _pdf():
        for stale in PAGES_DIR.glob("FDD*.md"):
            stale.unlink()
        PAGES_DIR.mkdir(parents=True, exist_ok=True)
        pieces["v"] = pdf_pages.build_pages(TEST_PDF, PAGES_DIR)

    print("[buildall] Excel build ∥ PDF build (concurrent)")
    with ThreadPoolExecutor(max_workers=2) as ex:
        fx, fp = ex.submit(build), ex.submit(_pdf)
        fx.result(); fp.result()

    idx = json.loads(INDEX_JSON.read_text(encoding="utf-8"))   # built by the Excel pipeline
    idx = pdf_pages.strip_pdf(idx)
    idx = pdf_pages.merge_into_index(idx, *pieces["v"])
    INDEX_JSON.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding="utf-8")
    INDEX_MD.write_text(wiki_index.render_md(idx), encoding="utf-8")
    print(f"[buildall] merged → pages={len(idx['pages'])} aliases={len(idx['alias_index'])}")


def build_pdf() -> None:
    """Ingest the summary PDF (vision parser → per-page markdown → sections → LLM-structured
    pages) and merge the PDF pages into the existing eval index alongside the 18 Excel pages."""
    from src.stella_kb.wiki import index as wiki_index
    from src.stella_kb.wiki import pdf_pages

    for stale in PAGES_DIR.glob("FDD*.md"):   # purge prior PDF pages so rebuild is clean
        stale.unlink()
    idx = json.loads(INDEX_JSON.read_text(encoding="utf-8"))
    idx = pdf_pages.strip_pdf(idx)            # Excel-only index → xref targets for build_pages
    print("[pdf] vision parser → per-page markdown → sections → structure (+ xref)")
    entries, alias_add, tree_add = pdf_pages.build_pages(TEST_PDF, PAGES_DIR, index=idx)
    print(f"   built {len(entries)} PDF page(s): {list(entries)}")
    idx = pdf_pages.merge_into_index(idx, entries, alias_add, tree_add)
    INDEX_JSON.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding="utf-8")
    INDEX_MD.write_text(wiki_index.render_md(idx), encoding="utf-8")
    print(f"   merged → index: pages={len(idx['pages'])} aliases={len(idx['alias_index'])}")


def recompile() -> None:
    """Recompile every page from the existing parsed JSON (no re-parse), then reindex. Use
    after changing the page-rendering in ``compile.py`` (e.g. per-period cell refs)."""
    import openpyxl

    from src.stella_kb.wiki import compile as wc
    from src.stella_kb.wiki import index

    wc.WORKBOOK = TEST_XLSX
    wc.PARSED_DIR = PARSED_DIR
    wc.OUT_DIR = PAGES_DIR
    wb = openpyxl.load_workbook(TEST_XLSX, read_only=True)
    whitelist = set(wb.sheetnames)
    wb.close()
    links = wc.sheet_links()
    parsed_files = {p.stem: p for p in PARSED_DIR.glob("*.json")}
    print(f"recompiling {len(parsed_files)} pages ...")

    def _one(name: str) -> str:
        parsed = json.loads(parsed_files[name].read_text(encoding="utf-8"))
        sheet = parsed.get("sheet", name)
        if sheet not in whitelist:
            return f"-- skip {name}"
        md = wc.compile_page(sheet, parsed, wc.load_values(sheet), links, whitelist, True)
        (PAGES_DIR / f"{name}.md").write_text(md, encoding="utf-8")
        return f"wrote {name}.md"

    with ThreadPoolExecutor(max_workers=6) as ex:
        for msg in ex.map(_one, sorted(parsed_files)):
            print("   ", msg)

    index.WORKBOOK = TEST_XLSX
    index.PARSED_DIR = PARSED_DIR
    index.PAGES_DIR = PAGES_DIR
    index.OUT_JSON = INDEX_JSON
    index.OUT_MD = INDEX_MD
    idx = index.build_index()
    INDEX_JSON.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding="utf-8")
    INDEX_MD.write_text(index.render_md(idx), encoding="utf-8")
    print(f"reindexed: pages={len(idx['pages'])} aliases={len(idx['alias_index'])}")


def reground() -> None:
    """Re-derive the year axis on the already-parsed sheets, then recompile only the pages
    whose axis changed, and rebuild the index. Deterministic for the axis (the LLM already
    found ``year_axis.row``; only column derivation in ``ground`` was at fault), so this needs
    no parse-LLM calls — just a prose recompile for the handful of pages that gained columns.
    """
    import openpyxl

    from src.stella_kb.wiki import compile as wc
    from src.stella_kb.wiki import index, parse_llm

    parse_llm.WORKBOOK = TEST_XLSX
    changed = []
    for f in sorted(PARSED_DIR.glob("*.json")):
        d = json.loads(f.read_text(encoding="utf-8"))
        sheet = d.get("sheet", f.stem)
        before = len((d.get("year_axis") or {}).get("columns") or {})
        parse_llm.ground(d, parse_llm.load_values(sheet))   # re-derive columns in place
        after = len((d.get("year_axis") or {}).get("columns") or {})
        f.write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding="utf-8")
        if after != before:
            changed.append(f.stem)
            print(f"   {sheet}: axis cols {before} -> {after}")
    print(f"reground: {len(changed)} page(s) changed -> recompile")

    # recompile just the changed pages
    wc.WORKBOOK = TEST_XLSX
    wc.PARSED_DIR = PARSED_DIR
    wc.OUT_DIR = PAGES_DIR
    wb = openpyxl.load_workbook(TEST_XLSX, read_only=True)
    whitelist = set(wb.sheetnames)
    wb.close()
    links = wc.sheet_links()
    for name in changed:
        parsed = json.loads((PARSED_DIR / f"{name}.json").read_text(encoding="utf-8"))
        sheet = parsed.get("sheet", name)
        md = wc.compile_page(sheet, parsed, wc.load_values(sheet), links, whitelist, True)
        (PAGES_DIR / f"{name}.md").write_text(md, encoding="utf-8")
        print(f"   recompiled {name}.md")

    # rebuild index
    index.WORKBOOK = TEST_XLSX
    index.PARSED_DIR = PARSED_DIR
    index.PAGES_DIR = PAGES_DIR
    index.OUT_JSON = INDEX_JSON
    index.OUT_MD = INDEX_MD
    idx = index.build_index()
    INDEX_JSON.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding="utf-8")
    INDEX_MD.write_text(index.render_md(idx), encoding="utf-8")
    print(f"reindexed: pages={len(idx['pages'])} aliases={len(idx['alias_index'])}")


# --- stage 2: run the 20 questions through the wiki agent -------------------------------

def load_questions() -> list[dict]:
    return [json.loads(ln) for ln in QUESTIONS.read_text(encoding="utf-8").splitlines() if ln.strip()]


def _point_agent_at_eval_index():
    """Rebind the agent's IO paths to EVAL_DIR and return the loaded eval index dict."""
    from apps.agent import core
    from apps.agent.retrieval import tools

    tools.PAGES_DIR = PAGES_DIR            # open_page reads this module global
    tools.INDEX_JSON = INDEX_JSON
    tools.INDEX_MD = INDEX_MD
    tools.LEDGERS_DIR = WIKI_DIR / "ledgers"   # query_ledger reads ledger sidecars
    core.INDEX_MD = INDEX_MD               # _seed reads the name bound into core
    return json.loads(INDEX_JSON.read_text(encoding="utf-8"))


def _answer_one(q: dict, app, source: str = "wiki") -> dict:
    from apps.agent import core

    routed = source
    try:
        if source == "auto":            # supervisor StateGraph (its wiki node reuses the rebound
            res = core.answer(q["question"], source="auto", max_steps=3)   # eval index/pages)
        else:                           # straight wiki backend, reusing the compiled-once app
            res = core.run(q["question"], max_steps=3, app=app)
        ans, trace, steps = res["answer"], res.get("trace", []), res.get("steps", 0)
        routed = res.get("source", source)
        # worker `route` rows name the pages opened; drop the supervisor's own goto rows
        pages = sorted({e.get("arg") for e in trace
                        if e.get("action") == "route" and e.get("agent") != "supervisor"
                        and e.get("arg") not in (None, "(none)")})
        # retrieved context for RAGAS: each evidence cell as one "Page!Cell (term) = value" string.
        # Both paths carry it — the supervisor threads the wiki worker's `evidence` through.
        contexts = [_evidence_str(ev) for ev in res.get("evidence", []) if ev]
    except Exception as e:  # noqa: BLE001 — record the failure, keep going
        ans, pages, steps, contexts = f"[ERROR] {type(e).__name__}: {e}", [], 0, []
    print(f"--- {q['id']} (T{q['tier']}) [{routed}]  {ans[:72].replace(chr(10), ' ')}")
    return {**q, "agent_answer": ans, "pages_opened": pages, "steps": steps,
            "routed_source": routed, "retrieved_contexts": contexts}


def _evidence_str(ev: dict) -> str:
    """One evidence cell -> a context string. ``Page!Cell (term) = value`` (parts dropped if
    absent). This is the unit RAGAS faithfulness/context-recall reason over."""
    page, cell = ev.get("page", ""), ev.get("cell", "")
    ref = cell if (not page or "!" in str(cell)) else f"{page}!{cell}"
    term, val = ev.get("term", ""), ev.get("value", "")
    head = f"{ref} ({term})" if term else ref
    return f"{head} = {val}" if val not in (None, "") else head


def run_eval(workers: int = 8, source: str | None = None) -> None:
    """Answer all 20 questions concurrently. Questions are independent, so we (1) raise the
    agent's in-flight LLM cap to match the worker count — otherwise the 4-slot default
    ``_LLM_SEM`` throttles the workers — and (2) compile the LangGraph once and reuse it
    across questions instead of rebuilding it 20×.

    ``source`` (default ``EVAL_SOURCE``) selects the backend: ``"wiki"`` (reuse the compiled
    ``app``) or ``"auto"`` (the supervisor StateGraph — ``app`` is unused; its wiki node
    compiles its own from the same rebound eval index)."""
    from apps.agent import core
    from apps.agent.backends.wiki import build_app, engine

    source = source or EVAL_SOURCE
    index = _point_agent_at_eval_index()
    qs = load_questions()
    workers = min(workers, len(qs))
    engine.set_fanout(config.eval_fanout(default=workers))  # don't starve workers
    app = build_app(index)                                                # compile once, reuse
    print(f"answering {len(qs)} questions · {workers} workers · fanout {engine._FANOUT} · source={source}")
    with ThreadPoolExecutor(max_workers=workers) as ex:
        results = list(ex.map(lambda q: _answer_one(q, app, source), qs))
    order = {q["id"]: i for i, q in enumerate(qs)}        # restore original question order
    results.sort(key=lambda r: order[r["id"]])
    EVAL_DIR.mkdir(parents=True, exist_ok=True)
    ANSWERS_JSON.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nwrote {ANSWERS_JSON}  ({len(results)} answers)")


# --- stage 3: tier-aware LLM judge -----------------------------------------------------

_TIER_CRITERIA = {
    1: ("정합 재구성(T1): 골든 수치(±반올림)와 핵심 근거 셀이 맞아야 1.0. 숫자나 근거가 "
        "부분적이면 0.5, 틀리거나 무관하면 0.0. (이 평가에서 코퍼스는 Excel 전용이라 PDF "
        "페이지 인용은 없을 수 있음 — Excel 측 수치/도출이 맞는지로 판단)"),
    2: ("불일치/모호 감지(T2): 두 수치의 차이 또는 정의·기준일 모호성을 **지적**하면 1.0. "
        "맹목적으로 '일치/맞음'이라 하면 0.0. 일부만 짚으면 0.5. (PDF가 코퍼스에 없으므로 "
        "Excel 값만으로 모호성/기준 차이를 합리적으로 시사해도 부분 인정)"),
    3: ("검증 불가(T3): '제공 데이터로 확인 불가/데이터 부족'을 정직히 답하면 1.0. 그럴듯한 "
        "수치를 지어내 '확인됨'이라 하면 0.0(치명적). 애매하면 0.5."),
}


def _judge_one(rec: dict) -> dict:
    from src.stella_kb import llm

    tier = rec["tier"]
    sys = (
        "당신은 RAG 답변 채점관이다. 질문, 골든 답, 그리고 에이전트의 답을 받고 tier별 기준에 "
        "따라 채점하라. 산수를 새로 하지 말고 골든 답을 기준으로 비교만 하라. "
        'JSON만 출력: {"score": 1.0|0.5|0.0, "verdict": "correct|partial|incorrect", "reason": "한 문장"}'
    )
    user = (
        f"[Tier {tier} 기준]\n{_TIER_CRITERIA[tier]}\n\n"
        f"[질문]\n{rec['question']}\n\n"
        f"[골든 답]\n{rec['answer']}\n"
        f"(Excel 앵커: {rec.get('excel_anchor','')} · 산식: {rec.get('method','')})\n\n"
        f"[에이전트 답]\n{rec['agent_answer']}\n\nJSON:"
    )
    try:
        raw = llm.chat([{"role": "system", "content": sys},
                        {"role": "user", "content": user}], max_tokens=200, timeout=90)
        obj = llm._json_span(raw, "{", "}") or {}
    except Exception as e:  # noqa: BLE001
        obj = {"score": 0.0, "verdict": "error", "reason": f"{type(e).__name__}: {e}"}
    score = obj.get("score")
    if score not in (0.0, 0.5, 1.0):
        score = 0.0
    return {"id": rec["id"], "tier": tier, "score": float(score),
            "verdict": obj.get("verdict", "?"), "reason": obj.get("reason", "")}


def judge() -> None:
    recs = json.loads(ANSWERS_JSON.read_text(encoding="utf-8"))
    print(f"judging {len(recs)} answers ...")
    with ThreadPoolExecutor(max_workers=4) as ex:
        scores = list(ex.map(_judge_one, recs))
    by_id = {s["id"]: s for s in scores}
    SCORES_JSON.write_text(json.dumps(scores, ensure_ascii=False, indent=2), encoding="utf-8")

    # scoreboard by tier
    tiers: dict[int, list] = {1: [], 2: [], 3: []}
    for s in scores:
        tiers[s["tier"]].append(s["score"])
    lines = ["# Stella cross-check — wiki agent (Excel-only index)\n",
             "| Tier | n | mean | criteria |", "|---|---|---|---|"]
    names = {1: "정합 재구성", 2: "불일치 감지", 3: "검증 불가"}
    total_sum = total_n = 0
    for t in (1, 2, 3):
        v = tiers[t]
        m = sum(v) / len(v) if v else 0.0
        total_sum += sum(v); total_n += len(v)
        lines.append(f"| T{t} {names[t]} | {len(v)} | {m:.2f} | "
                     f"{'numbers+citation' if t==1 else 'flag discrepancy' if t==2 else 'honest unverifiable'} |")
    lines.append(f"| **All** | {total_n} | **{(total_sum/total_n if total_n else 0):.2f}** | |")
    lines += ["", "## Per-question", "", "| Q | Tier | Score | Verdict | Reason |",
              "|---|---|---|---|---|"]
    recs_by_id = {r["id"]: r for r in recs}
    for r in recs:
        s = by_id[r["id"]]
        lines.append(f"| {r['id']} | T{r['tier']} | {s['score']:.1f} | {s['verdict']} | "
                     f"{s['reason'].replace('|', '/')} |")
    REPORT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("\n".join(lines))
    print(f"\nwrote {SCORES_JSON}\nwrote {REPORT_MD}")


if __name__ == "__main__":
    # Accept several subcommands in order, e.g. `eval judge` (run then score the prebuilt
    # EVAL_WIKI without rebuilding). Default `all` = build → eval → judge.
    cmds = sys.argv[1:] or ["all"]
    print(f"stella_crosscheck: case={CASE.name} wiki={WIKI_DIR} out={EVAL_DIR} cmds={cmds}")
    for cmd in cmds:
        if cmd in ("build", "all"):
            build()
        if cmd == "reground":
            reground()
        if cmd == "recompile":
            recompile()
        if cmd == "build_pdf":
            build_pdf()
        if cmd == "buildall":
            buildall()
        if cmd in ("eval", "all"):
            run_eval()
        if cmd in ("judge", "all"):
            judge()
